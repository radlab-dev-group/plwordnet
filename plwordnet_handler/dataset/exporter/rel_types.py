import logging

from typing import List, Optional, Dict
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from plwordnet_handler.base.structure.polishwordnet import PolishWordnet
from plwordnet_handler.base.structure.elems.rel_type import RelationType
from plwordnet_handler.base.connectors.db.db_connector import (
    PlWordnetAPIMySQLDbConnector,
)


class RelationTypesExporter:
    """
    Class for exporting relation types to Excel file.
    """

    HEADERS = [
        "ID",
        "PARENT_ID",
        "REVERSE_ID",
        "autoreverse",
        "posstr",
        "pwn",
        "name",
        "shortcut",
        "description",
        "LU_relations_count",
        "Synset_relations_count",
    ]

    def __init__(self, connector: PlWordnetAPIMySQLDbConnector):
        """
        Initialize the exporter with a database connector.

        Args:
            connector: Database connector for accessing plWordnet data
        """
        self.connector = connector
        self.logger = logging.getLogger(__name__)

    def get_relation_types(
        self, limit: Optional[int] = None
    ) -> Optional[List[RelationType]]:
        """
        Get relation types from the database.

        Args:
            limit: Optional limit for the number of results

        Returns:
            List of relation types or None if error occurred
        """
        try:
            with PolishWordnet(connector=self.connector) as pl_wn:
                return pl_wn.get_relation_types(limit=limit)
        except Exception as e:
            self.logger.error(f"Error retrieving relation types: {e}")
            return None

    def get_lexical_relation_counts(
        self, limit: Optional[int] = None
    ) -> Dict[int, int]:
        """
        Get count of lexical relations by relation type ID.

        Args:
            limit: Optional limit for the number of results

        Returns:
            Dictionary mapping REL_ID to count of lexical relations
        """
        try:
            with PolishWordnet(connector=self.connector) as pl_wn:
                lexical_relations = pl_wn.get_lexical_relations(limit=limit)
                if not lexical_relations:
                    return {}

                relation_counts = defaultdict(int)
                for relation in lexical_relations:
                    relation_counts[relation.REL_ID] += 1

                return dict(relation_counts)
        except Exception as e:
            self.logger.error(f"Error retrieving lexical relations: {e}")
            return {}

    def get_synset_relation_counts(
        self, limit: Optional[int] = None
    ) -> Dict[int, int]:
        """
        Get count of synset relations by relation type ID.

        Args:
            limit: Optional limit for the number of results

        Returns:
            Dictionary mapping REL_ID to count of synset relations
        """
        try:
            with PolishWordnet(connector=self.connector) as pl_wn:
                synset_relations = pl_wn.get_synset_relations(limit=limit)
                if not synset_relations:
                    return {}

                relation_counts = defaultdict(int)
                for relation in synset_relations:
                    relation_counts[relation.REL_ID] += 1

                return dict(relation_counts)
        except Exception as e:
            self.logger.error(f"Error retrieving synset relations: {e}")
            return {}

    def export_to_xlsx(self, output_file: str, limit: Optional[int] = None):
        """
        Export relation types to an Excel file.

        Args:
            output_file: Path to the output Excel file
            limit: Optional limit for the number of results

        Returns:
            Boolean indicating success or failure
        """
        self.logger.info(f"Starting export of relation types to {output_file}")

        relation_types = self.get_relation_types(limit=limit)
        if not relation_types:
            self.logger.error("No relation types data retrieved")
            return False

        self.logger.info("Retrieving lexical relation counts...")
        lu_relation_counts = self.get_lexical_relation_counts(limit=limit)

        self.logger.info("Retrieving synset relation counts...")
        synset_relation_counts = self.get_synset_relation_counts(limit=limit)

        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Relation Types"

            self._add_header(worksheet=worksheet)
            self._add_data(
                relation_types=relation_types,
                worksheet=worksheet,
                lu_counts=lu_relation_counts,
                synset_counts=synset_relation_counts,
            )
            self._adjust_column_widths(worksheet=worksheet)

            workbook.save(output_file)

            self.logger.info(
                f"Successfully exported {len(relation_types)} "
                f"relation types to {output_file}"
            )
            return True
        except Exception as e:
            self.logger.error(f"Error during Excel export: {e}")
            return False

    def _add_header(self, worksheet: Worksheet):
        """
        Add the header row to the worksheet with formatting.

        Creates a styled header row with white text on a blue background.

        Args:
            worksheet: The Excel worksheet to add headers to
        """

        self.logger.info(f"Adding header row to worksheet: {worksheet}")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        for col_idx, h_label in enumerate(self.HEADERS, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=h_label)
            cell.font = header_font
            cell.fill = header_fill

    def _add_data(
        self,
        relation_types: List[RelationType],
        worksheet: Worksheet,
        lu_counts: Dict[int, int],
        synset_counts: Dict[int, int],
    ):
        """
        Add relation types data to the worksheet.

        Populates the worksheet with relation type data starting from row 2.
        Each RelationType object's attributes are mapped to specific columns.

        Args:
            relation_types: List of RelationType objects to export
            worksheet: The Excel worksheet to populate with data
            lu_counts: Dictionary mapping REL_ID to count of lexical relations
            synset_counts: Dictionary mapping REL_ID to count of synset relations
        """
        self.logger.info(f"Adding data row to worksheet: {worksheet}")

        for row_idx, rel_type in enumerate(relation_types, 2):
            worksheet.cell(row=row_idx, column=1, value=rel_type.ID)
            worksheet.cell(row=row_idx, column=2, value=rel_type.PARENT_ID)
            worksheet.cell(row=row_idx, column=3, value=rel_type.REVERSE_ID)
            worksheet.cell(row=row_idx, column=4, value=rel_type.autoreverse)
            worksheet.cell(row=row_idx, column=5, value=rel_type.posstr)
            worksheet.cell(row=row_idx, column=6, value=rel_type.pwn)
            worksheet.cell(row=row_idx, column=7, value=rel_type.name)
            worksheet.cell(row=row_idx, column=8, value=rel_type.shortcut)
            worksheet.cell(row=row_idx, column=9, value=rel_type.description)
            worksheet.cell(
                row=row_idx, column=10, value=lu_counts.get(rel_type.ID, 0)
            )
            worksheet.cell(
                row=row_idx, column=11, value=synset_counts.get(rel_type.ID, 0)
            )

    def _adjust_column_widths(self, worksheet: Worksheet):
        """
        Automatically adjust column widths based on content length.

        Iterates through all columns and sets width to accommodate the longest
        content in each column, with a maximum width limit of 50 characters.

        Args:
            worksheet: The Excel worksheet to adjust column widths for
        """
        self.logger.info(f"Adjusting columns in worksheet: {worksheet}")

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
